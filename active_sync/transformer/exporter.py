"""Exportação opcional do DataFrame tratado para validação visual."""

from __future__ import annotations

import logging
from pathlib import Path
import tempfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from active_sync.exceptions import TransformationValidationError

from .columns import DATE_COLUMNS, NUMERIC_COLUMNS, OUTPUT_COLUMNS, TEXT_COLUMNS
from .validator import validate_output_dataframe


def export_validation_excel(
    frame: pd.DataFrame,
    output_path: Path | str = Path("performance_validacao.xlsx"),
    logger: logging.Logger | None = None,
) -> Path:
    """Gera um Excel simples para comparação visual com o Power Query."""
    validation = validate_output_dataframe(frame)
    if not validation.is_valid:
        raise TransformationValidationError("; ".join(validation.errors))

    destination = Path(output_path).resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", dir=destination.parent, delete=False
        ) as temporary:
            temporary_path = Path(temporary.name)
        frame.to_excel(
            temporary_path,
            index=False,
            sheet_name="Performance Entrega",
            engine="openpyxl",
        )
        workbook = load_workbook(temporary_path)
        sheet = workbook["Performance Entrega"]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 24

        for index, column in enumerate(OUTPUT_COLUMNS, start=1):
            letter = get_column_letter(index)
            values = [sheet.cell(row=row, column=index).value for row in range(1, min(sheet.max_row, 101) + 1)]
            width = min(max(len(str(value)) if value is not None else 0 for value in values) + 2, 35)
            sheet.column_dimensions[letter].width = max(width, 12)
            if column in DATE_COLUMNS:
                for cell in sheet[letter][1:]:
                    cell.number_format = "dd/mm/yyyy"
            elif column in NUMERIC_COLUMNS:
                for cell in sheet[letter][1:]:
                    cell.number_format = "#,##0.00"
            elif column in TEXT_COLUMNS:
                for cell in sheet[letter][1:]:
                    cell.number_format = "@"
        workbook.save(temporary_path)
        temporary_path.replace(destination)
        temporary_path = None
    finally:
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()

    if logger:
        logger.info("Excel de validação exportado: %s", destination)
    return destination
