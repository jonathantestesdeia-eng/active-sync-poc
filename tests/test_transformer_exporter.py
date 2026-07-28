from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from active_sync.transformer.columns import OUTPUT_COLUMNS
from active_sync.transformer.exporter import export_validation_excel
from active_sync.transformer.transforms import transform_dataframe


def source_frame() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Destinatário": ["Cliente A"],
            "Cidade Origem": ["CAMBUI"],
            "Cidade Destino": ["TUPA"],
            "UF Destino": ["SP"],
            "Nota Fiscal": ["000123"],
            "CTe": ["000456"],
            "Valor Frete": ["123.45"],
            "Saída": ["21/07/2026"],
            "Previsão": ["29/07/2026"],
            "Entrega": [None],
            "Transportador": ["Transportadora X"],
            "Tipo": ["ENTREGA NORMAL"],
            "Observacao": [None],
            "Trecho": ["NORMAL"],
        }
    )


def test_exports_validation_excel_with_exact_headers(tmp_path: Path) -> None:
    frame = transform_dataframe(source_frame())

    output = export_validation_excel(frame, tmp_path / "performance_validacao.xlsx")

    workbook = load_workbook(output, read_only=False, data_only=True)
    sheet = workbook["Performance Entrega"]
    headers = tuple(cell.value for cell in sheet[1])
    assert headers == OUTPUT_COLUMNS
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == sheet.dimensions
    assert sheet.max_row == 2
    assert sheet.max_column == 22
